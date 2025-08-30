from __future__ import annotations

import os
import re
import io
import json
import time
import random
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime
from pathlib import Path
import boto3
from botocore.exceptions import ClientError
import requests
import pandas as pd


BEDROCK_REGION = os.getenv("BEDROCK_REGION", "us-east-1")
# Latest Claude 3.5 Sonnet on Bedrock (adjust if your account exposes a different alias/version)
BEDROCK_MODEL_ID = os.getenv("BEDROCK_MODEL_ID", "anthropic.claude-3-5-sonnet-20240620-v1:0")

# Global default sampling parameters (well-balanced for business tasks)
GLOBAL_PARAMS = {
    "temperature": float(os.getenv("BEDROCK_TEMPERATURE", 0.6)),
    "top_p": float(os.getenv("BEDROCK_TOP_P", 0.96)),
    "top_k": int(os.getenv("BEDROCK_TOP_K", 250)),
    "max_tokens": int(os.getenv("BEDROCK_MAX_TOKENS", 2400)),
    "stop_sequences": []  # you can add custom stops like ["</END>"]
}

# Retry configuration
RETRY_CONFIG = {
    "max_retries": int(os.getenv("BEDROCK_MAX_RETRIES", 5)),
    "base_delay": float(os.getenv("BEDROCK_BASE_DELAY", 1.0))
}

# Per-module parameter profiles (you can tweak live in the UI)
PARAM_PROFILES = {
    "persona":  {"temperature": 0.65, "top_p": 0.96, "top_k": 300, "max_tokens": 2600},
    "usp":      {"temperature": 0.5,  "top_p": 0.95, "top_k": 200, "max_tokens": 2000},
    "comp":     {"temperature": 0.4,  "top_p": 0.92, "top_k": 150, "max_tokens": 2000},
    "message":  {"temperature": 0.55, "top_p": 0.95, "top_k": 220, "max_tokens": 2600},
    "retention":{"temperature": 0.55, "top_p": 0.95, "top_k": 220, "max_tokens": 2200},
    "evidence": {"temperature": 0.4,  "top_p": 0.92, "top_k": 150, "max_tokens": 1800},
    "validate": {"temperature": 0.3,  "top_p": 0.9,  "top_k": 120, "max_tokens": 1200},
}

# -----------------------------

class SimpleObserver:
    def __init__(self):
        self.cloudwatch = None
        self.log_file = Path("app_metrics.jsonl")
        try:
            self.cloudwatch = boto3.client('cloudwatch', region_name=BEDROCK_REGION)
        except:
            pass
    
    def log_event(self, event_type: str, data: dict):
        log_entry = {
            "timestamp": datetime.now().isoformat(),
            "event_type": event_type,
            **data
        }
        try:
            with open(self.log_file, "a") as f:
                f.write(json.dumps(log_entry) + "\n")
        except:
            pass
    
    def log_bedrock_call(self, model_id: str, tokens: int, latency: float, success: bool):
        self.log_event("bedrock_call", {
            "model_id": model_id,
            "tokens": tokens,
            "latency_seconds": latency,
            "success": success
        })
        
        if self.cloudwatch:
            try:
                # Calculate estimated cost (Claude 3.5 Sonnet pricing)
                estimated_cost = tokens * 0.000015  # $15 per 1M output tokens (approximate)
                
                metrics = [
                    {'MetricName': 'BedrockCalls', 'Value': 1, 'Unit': 'Count'},
                    {'MetricName': 'TokensUsed', 'Value': tokens, 'Unit': 'Count'},
                    {'MetricName': 'Latency', 'Value': latency, 'Unit': 'Seconds'},
                    {'MetricName': 'EstimatedCost', 'Value': estimated_cost, 'Unit': 'None'}
                ]
                if not success:
                    metrics.append({'MetricName': 'Errors', 'Value': 1, 'Unit': 'Count'})
                
                # Add dimensions for better filtering
                for metric in metrics:
                    metric['Dimensions'] = [{'Name': 'ModelId', 'Value': model_id}]
                
                self.cloudwatch.put_metric_data(Namespace="ZeroPitchForge", MetricData=metrics)
            except:
                pass
    
    def log_custom_metric(self, metric_name: str, value: float, unit: str = 'Count', dimensions: dict = None):
        """Log any custom metric to CloudWatch"""
        if self.cloudwatch:
            try:
                metric_data = {
                    'MetricName': metric_name,
                    'Value': value,
                    'Unit': unit
                }
                if dimensions:
                    metric_data['Dimensions'] = [{'Name': k, 'Value': v} for k, v in dimensions.items()]
                
                self.cloudwatch.put_metric_data(Namespace="ZeroPitchForge", MetricData=[metric_data])
            except:
                pass

_observer = SimpleObserver()

# -----------------------------
# Bedrock client & helpers
# -----------------------------

def _bedrock_client():
    return boto3.client("bedrock-runtime", region_name=BEDROCK_REGION)

def _merge_params(base: Dict[str, Any], override: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    merged = {**base}
    if override:
        for k, v in override.items():
            if v is not None:
                merged[k] = v
    return merged

def _retry_with_backoff(func, max_retries: int = 5, base_delay: float = 1.0):
    """Retry function with exponential backoff for throttling errors."""
    for attempt in range(max_retries):
        try:
            return func()
        except ClientError as e:
            error_code = e.response.get('Error', {}).get('Code', '')
            if error_code == 'ThrottlingException' and attempt < max_retries - 1:
                delay = base_delay * (2 ** attempt) + random.uniform(0, 1)
                time.sleep(delay)
                continue
            raise e
        except Exception as e:
            if "ThrottlingException" in str(e) and attempt < max_retries - 1:
                delay = base_delay * (2 ** attempt) + random.uniform(0, 1)
                time.sleep(delay)
                continue
            raise e

def invoke_claude_messages(
    user_text: str,
    system_text: Optional[str] = None,
    params: Optional[Dict[str, Any]] = None,
    model_id: Optional[str] = None
) -> str:
    """
    Calls Claude 3.x Messages API on Bedrock with retry logic for throttling errors.
    """
    model_id = model_id or BEDROCK_MODEL_ID
    p = _merge_params(GLOBAL_PARAMS, params)
    start_time = time.time()

    payload = {
        "anthropic_version": "bedrock-2023-05-31",
        "max_tokens": p["max_tokens"],
        "temperature": p["temperature"],
        "top_p": p["top_p"],
        "top_k": p["top_k"],
        "messages": [
            {
                "role": "user",
                "content": [{"type": "text", "text": user_text}]
            }
        ]
    }
    if system_text:
        payload["system"] = system_text
    if p.get("stop_sequences"):
        payload["stop_sequences"] = p["stop_sequences"]

    client = _bedrock_client()
    
    def _invoke():
        resp = client.invoke_model(
            modelId=model_id,
            contentType="application/json",
            accept="application/json",
            body=json.dumps(payload)
        )
        body = json.loads(resp["body"].read())
        content = body.get("content", [])
        return content[0].get("text", "") if content else ""
    
    try:
        result = _retry_with_backoff(_invoke, RETRY_CONFIG["max_retries"], RETRY_CONFIG["base_delay"])
        latency = time.time() - start_time
        _observer.log_bedrock_call(model_id, len(result), latency, True)
        return result
    except Exception as e:
        latency = time.time() - start_time
        _observer.log_bedrock_call(model_id, 0, latency, False)
        raise

def parse_json_safely(text: str) -> Any:
    """
    Attempts to parse JSON from the LLM output. If it fails,
    tries to extract the largest JSON object/array via regex.
    """
    text = text.strip()
    # Direct attempt
    try:
        return json.loads(text)
    except Exception:
        pass
    # Try to extract a fenced JSON block ```json ... ```
    m = re.search(r"```json(.*?)```", text, flags=re.DOTALL | re.IGNORECASE)
    if m:
        snippet = m.group(1).strip()
        try:
            return json.loads(snippet)
        except Exception:
            pass
    # Try to find a top-level array/object
    m2 = re.search(r"(\{.*\}|\[.*\])", text, flags=re.DOTALL)

    if m2:
        try:
            return json.loads(m2.group(1))
        except Exception:
            pass
    # Fallback: return raw text
    return {"raw_text": text}

# -----------------------------
# Prompt builders (guardrailed)
# -----------------------------

SYS_PERSONA = (
    "You are a precise, business-savvy persona generator. "
    "You write concrete, non-generic, testable details. "
    "Never include chain-of-thought. Output ONLY valid JSON."
)

def prompt_persona(business_desc: str, categories: List[str], regions: List[str], tone: str = "practical") -> str:
    cats = ", ".join(categories) if categories else "general"
    regs = ", ".join(regions) if regions else "global"
    return f"""
Generate 3 distinct narrative buyer personas for the business below. Be specific, non-generic, and align to region/channel norms.

Business description: {business_desc}
Product/service categories: {cats}
Target regions: {regs}
Preferred tone: {tone}

Return ONLY a JSON array of 3 objects with fields:
- "persona_id": short stable id (e.g., "ops_lead_indie")
- "name": short label
- "segment": role/vertical context
- "trigger_moment": concrete situation that spikes the pain
- "emotional_stakes": concise human stakes (fear, hope)
- "jobs_to_be_done": array of functional + emotional jobs
- "pain_points": array of concrete pains (avoid generic)
- "objections": array of specific objections you might hear
- "channels": array of discovery/trust channels (2–4)
- "budget_band": one of ["<$100/mo","$100-$499/mo","$500-$1,999/mo","$2,000+/mo"]
- "regions": array of region tags
- "sample_quote": one-line voice-of-customer

Constraints:
- Do not fabricate personal data.
- Keep each field concise but information-dense.
- Ensure personas are distinct in segment and psychographics.
"""

SYS_USP = (
    "You are a crisp USP distiller. You turn features into sharp claims with proofs. "
    "No chain-of-thought. Output ONLY valid JSON."
)

def prompt_usp(features: List[str], benefits: List[str], category_norms: Optional[List[str]] = None) -> str:
    feats = "- " + "\n- ".join(features) if features else "-"
    bens = "- " + "\n- ".join(benefits) if benefits else "-"
    norms = "- " + "\n- ".join(category_norms) if category_norms else "-"

    return f"""
Given product inputs, produce a USP grid with claims, proofs, and counter-claims. Be specific and testable.

Features:
{feats}

Benefits:
{bens}

Category norms (what buyers expect or jargon to avoid):
{norms}

Return ONLY a JSON object:
{{
  "table_stakes": ["..."],
  "differentiators": ["..."],
  "reasons_to_believe": [{{"claim":"...", "proof":"demo, metric, guarantee, certification"}}],
  "counter_claims": [{{"competitor_claim":"...", "your_rejoinder":"..."}}]
}}
"""

SYS_COMP = (
    "You are a competitive angle miner. You summarize public snippets into positioning patterns and whitespace. "
    "No chain-of-thought. Output ONLY valid JSON."
)

def prompt_competitive(snippets: List[Dict[str, str]], max_snippets: int = 6) -> str:
    joined = "\n\n".join([f"Source: {s.get('url','')}\nSnippet:\n{s.get('text','')[:1200]}" for s in snippets[:max_snippets]])
    return f"""
Using the public snippets below, extract competitor claims, pricing patterns, and messaging themes.
Return whitespace opportunities that avoid me-too phrasing.

Public snippets:
{joined}

Return ONLY a JSON object:
{{
  "claims": [{{"statement":"...", "source_url":"...", "confidence":"high|med|low"}}],
  "pricing_patterns": ["..."],
  "messaging_themes": ["..."],
  "whitespace_opportunities": ["..."]
}}
"""

SYS_MESSAGE = (
    "You are a message house synthesizer. You craft a core promise, proof pillars, and on-brand copy. "
    "No chain-of-thought. Output ONLY valid JSON."
)

def prompt_message_house(personas_json: str, usp_json: str) -> str:
    return f"""
Create a concise message house and copy assets using the personas and USP grid.

Personas (JSON):
{personas_json}

USP grid (JSON):
{usp_json}

Return ONLY a JSON object:
{{
  "core_promise": "one-sentence specific benefit (emotion + outcome)",
  "pillars": [
    {{"claim":"...", "proof":"...", "example":"..."}},
    {{"claim":"...", "proof":"...", "example":"..."}},
    {{"claim":"...", "proof":"...", "example":"..."}}
  ],
  "taglines": ["7 concise options across tones"],
  "landing": {{
    "headline": "...",
    "subhead": "...",
    "cta": "action verb + outcome",
    "social_proof_slot": "what belongs here"
  }},
  "headlines": ["6 paid/organic headline options"],
  "objection_handling": [
    {{"claim":"...", "concern":"...", "response":"...", "proof":"..."}}
  ]
}}
"""

SYS_RETENTION = (
    "You are a retention and activation planner. You map first-run steps and write a short email sequence. "
    "No chain-of-thought. Output ONLY valid JSON."
)

def prompt_retention(personas_json: str, objection_themes: Optional[List[str]] = None) -> str:
    obs = "- " + "\n- ".join(objection_themes) if objection_themes else "-"
    return f"""
Design an activation map and a 5-email sequence aligned to persona pains and objections.

Key objections to address:
{obs}

Personas (JSON):
{personas_json}

Return ONLY a JSON object:
{{
  "activation_map": {{
    "first_run_checklist": ["..."],
    "aha_moments": ["..."],
    "nudges": ["..."]
  }},
  "email_sequence": [
    {{"day": 0, "subject": "...", "goal": "...", "body_outline": "...", "cta": "..."}},
    {{"day": 2, "subject": "...", "goal": "...", "body_outline": "...", "cta": "..."}},
    {{"day": 5, "subject": "...", "goal": "...", "body_outline": "...", "cta": "..."}},
    {{"day": 7, "subject": "...", "goal": "...", "body_outline": "...", "cta": "..."}},
    {{"day": 14, "subject": "...", "goal": "...", "body_outline": "...", "cta": "..."}}
  ]
}}
"""

SYS_EVIDENCE = (
    "You are an evidence summarizer. You compress public snippets into signals with source and confidence. "
    "No chain-of-thought. Output ONLY valid JSON."
)

def prompt_evidence(snippets: List[Dict[str, str]], business_desc: str, categories: List[str]) -> str:
    joined = "\n\n".join([f"Source: {s.get('url','')}\nSnippet:\n{s.get('text','')[:1200]}" for s in snippets])
    cats = ", ".join(categories) if categories else "general"
    return f"""
Summarize external evidence relevant to the business and categories.

Business: {business_desc}
Categories: {cats}

Public snippets:
{joined}

Return ONLY a JSON object:
{{
  "signals": [{{"statement":"...", "source_url":"...", "confidence":"high|med|low"}}],
  "implications": ["..."],
  "conflicts": ["..."]
}}
"""

# -----------------------------
# Internet-fed evidence (simple)
# -----------------------------

def fetch_public_text(url: str, timeout: int = 12) -> str:
    try:
        headers = {"User-Agent": "ZeroDataBot/1.0 (+research; contact: none)"}
        r = requests.get(url, timeout=timeout, headers=headers)
        r.raise_for_status()
        html = r.text
        # Minimal HTML -> text; for higher fidelity, integrate trafilatura/bs4
        text = re.sub(r"<script[\s\S]*?</script>", " ", html, flags=re.I)
        text = re.sub(r"<style[\s\S]*?</style>", " ", text, flags=re.I)
        text = re.sub(r"<[^>]+>", " ", text)
        text = re.sub(r"\s+", " ", text).strip()
        return text[:12000]
    except Exception:
        return ""

def gather_snippets(urls: List[str]) -> List[Dict[str, str]]:
    items = []
    for u in urls:
        t = fetch_public_text(u)
        if len(t) > 500:
            items.append({"url": u, "text": t})
    return items

# -----------------------------
# Orchestration helpers
# -----------------------------

def run_personas(business_desc: str, categories: List[str], regions: List[str], tone: str,
                 param_override: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
    txt = invoke_claude_messages(
        user_text=prompt_persona(business_desc, categories, regions, tone),
        system_text=SYS_PERSONA,
        params=_merge_params(PARAM_PROFILES["persona"], param_override)
    )
    return parse_json_safely(txt)

def run_usp(features: List[str], benefits: List[str], category_norms: Optional[List[str]] = None,
            param_override: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    txt = invoke_claude_messages(
        user_text=prompt_usp(features, benefits, category_norms),
        system_text=SYS_USP,
        params=_merge_params(PARAM_PROFILES["usp"], param_override)
    )
    return parse_json_safely(txt)

def run_competitive(snippets: List[Dict[str, str]],
                    param_override: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    txt = invoke_claude_messages(
        user_text=prompt_competitive(snippets),
        system_text=SYS_COMP,
        params=_merge_params(PARAM_PROFILES["comp"], param_override)
    )
    return parse_json_safely(txt)

def run_message_house(personas: List[Dict[str, Any]], usp: Dict[str, Any],
                      param_override: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    txt = invoke_claude_messages(
        user_text=prompt_message_house(json.dumps(personas, ensure_ascii=False), json.dumps(usp, ensure_ascii=False)),
        system_text=SYS_MESSAGE,
        params=_merge_params(PARAM_PROFILES["message"], param_override)
    )
    return parse_json_safely(txt)

def run_retention(personas: List[Dict[str, Any]], objection_themes: Optional[List[str]] = None,
                  param_override: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    txt = invoke_claude_messages(
        user_text=prompt_retention(json.dumps(personas, ensure_ascii=False), objection_themes),
        system_text=SYS_RETENTION,
        params=_merge_params(PARAM_PROFILES["retention"], param_override)
    )
    return parse_json_safely(txt)

def run_evidence(business_desc: str, categories: List[str], urls: List[str],
                 param_override: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    snips = gather_snippets(urls)
    if not snips:
        return {"signals": [], "implications": [], "conflicts": []}
    txt = invoke_claude_messages(
        user_text=prompt_evidence(snips, business_desc, categories),
        system_text=SYS_EVIDENCE,
        params=_merge_params(PARAM_PROFILES["evidence"], param_override)
    )
    return parse_json_safely(txt)

# -----------------------------
# Excel export (DAX-ready)
# -----------------------------

def _persona_df(personas: List[Dict[str, Any]]) -> pd.DataFrame:
    df = pd.DataFrame(personas)
    # Normalize some fields
    for col in ["jobs_to_be_done", "pain_points", "objections", "channels", "regions"]:
        if col in df.columns:
            df[col] = df[col].apply(lambda v: ", ".join(v) if isinstance(v, list) else v)
    # DAX-friendly helper columns
    df["IsHighRisk"] = df.get("objections", "").apply(
        lambda x: 1 if isinstance(x, str) and any(term in x.lower() for term in ["security","compliance","migration","vendor","lock-in"]) else 0
    )
    df["ChannelFitLinkedIn"] = df.get("channels", "").apply(lambda x: 1 if isinstance(x, str) and "linkedin" in x.lower() else 0)
    df["ChannelFitEmail"] = df.get("channels", "").apply(lambda x: 1 if isinstance(x, str) and "email" in x.lower() else 0)

    def score(row):
        score = 0
        pains = (row.get("pain_points") or "").lower()
        if "time" in pains: score += 20
        if "cost" in pains: score += 20
        bb = (row.get("budget_band") or "")
        if "$500" in bb or "$2,000" in bb: score += 20
        if row.get("ChannelFitLinkedIn"): score += 20
        if row.get("ChannelFitEmail"): score += 20
        return score
    df["PersonaScore"] = df.apply(score, axis=1)
    return df

# def _dictlist_to_df(name: str, obj: Any) -> pd.DataFrame:
#     if isinstance(obj, list):
#         # Handle list of dicts vs list of scalars
#         if all(isinstance(x, dict) for x in obj):
#             return pd.json_normalize(obj)
#         else:
#             # Turn list of scalars into a single-column DataFrame
#             return pd.DataFrame({name: obj})
#     if isinstance(obj, dict):
#         return pd.json_normalize(obj)
#     return pd.DataFrame([{name: obj}])

def _dictlist_to_df(name: str, obj: Any) -> pd.DataFrame:
    if isinstance(obj, list):
        if all(isinstance(x, dict) for x in obj):
            df = pd.json_normalize(obj)
        else:
            df = pd.DataFrame({name: obj})
    elif isinstance(obj, dict):
        df = pd.json_normalize(obj)
    else:
        df = pd.DataFrame([{name: obj}])

    # 🔑 Convert any nested lists/dicts inside cells into JSON strings
    for col in df.columns:
        df[col] = df[col].apply(
            lambda v: json.dumps(v, ensure_ascii=False) if isinstance(v, (list, dict)) else v
        )
    return df

def export_excel(personas: List[Dict[str, Any]],
                 usp: Dict[str, Any],
                 comp: Dict[str, Any],
                 message_house: Dict[str, Any],
                 retention: Dict[str, Any]) -> bytes:
    """
    Returns an in-memory Excel file bytes for download. Sheets include DAX-friendly columns.
    """
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()
    # Personas
    ws = wb.active
    ws.title = "Personas"
    dfp = _persona_df(personas)
    for r in dataframe_to_rows(dfp, index=False, header=True):
        ws.append(r)

    # USP
    ws2 = wb.create_sheet("USP")
    dfu = _dictlist_to_df("usp", usp)
    for r in dataframe_to_rows(dfu, index=False, header=True):
        ws2.append(r)

    # Competitive insights
    ws3 = wb.create_sheet("Competitive")
    dfc = _dictlist_to_df("comp", comp)
    for r in dataframe_to_rows(dfc, index=False, header=True):
        ws3.append(r)

    # Message House
    ws4 = wb.create_sheet("MessageHouse")
    dfm = _dictlist_to_df("message_house", message_house)
    for r in dataframe_to_rows(dfm, index=False, header=True):
        ws4.append(r)

    # Retention
    ws5 = wb.create_sheet("Retention")
    dfr = _dictlist_to_df("retention", retention)
    for r in dataframe_to_rows(dfr, index=False, header=True):
        ws5.append(r)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()

    return bio.read()